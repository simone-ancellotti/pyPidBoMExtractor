from .utils import isTagBlock, getTagCode, findTypeBlockFromTag
from .extractor import extract_blocks_with_attributes_and_dimensions
import openpyxl
from openpyxl.styles import PatternFill
import os
import re

# Mapping between BOM keys and Excel column headers
header_mapping = {
        'count': '#',
        'targetObjectType': 'L',
        'targetObjectLoopNumber': 'N',
        'targetObjectType2nd': 'D',
        #'TYPE': 'Type',
        #'DESCRIPTION': 'Description',
        #'CONNECTIONTYPE':'C type'
    }

tags_xls2dxf= {'P&ID TAG':'P&ID TAG','Type':'TYPE', 'C type':'CONNECTIONTYPE', 'Description':'DESCRIPTION',
               'Fluid':'FLUID', 'Unit':'UNIT', 'Skid':'SKID', 'Type':'TYPE', 
               'Material':'MATERIAL', 'Seal Mat.':'SEAL_MAT', 
               'P     (kW)':'POWER_KW','PN   (bar)':'PN_bar', 
               'Act NO/NC':'ACT_NO_NC', 'Size':'SIZE','cap. (tanks L)':'CAP.(TANK L)', 
               'Q (m3/h)':'Q(m3/h)', 'Supplier':'SUPPLIER', 'Brand':'BRAND',
               'Model':'MODEL', 'Notes':'NOTES', 'Datasheet':'DATASHEET'
               }
tags_dxf2xls = {value: key for key, value in tags_xls2dxf.items()}
header_mapping.update(tags_dxf2xls)
header_mapping_reverse = {value: key for key, value in header_mapping.items()}


def filterBOM_Ignore(bom,tagToIgnore,valuesToIgnore):
    bom_filtered=bom.copy()
    for key in bom.keys():
        component = bom[key]
        tagKey = component.get(tagToIgnore)
        if tagKey:
            if (component[tagToIgnore] in valuesToIgnore):
                del bom_filtered[key]
    return bom_filtered
                
def stripField(field):
    if isinstance(field, str):
        field = field.strip() or None
    return field

def generate_bom(components):
    # feel free to change the logic
    tagBlocks = [c for c in components if isTagBlock(c)]
    bom = {}
    blocks_notValid = {}
    i=-1
    for component in tagBlocks:
        i+=1
        number = i+1
        #block_name = component['block_name']
        attributes = component['attributes']
        #contains_circle = component['contains_circle']
        
        #print(component)
        # extraction of Tags
        targetObjectExtracted = getTagCode(component)
        if targetObjectExtracted:
            targetObjectType        = targetObjectExtracted.get('targetObjectType')
            targetObjectLoopNumber  = targetObjectExtracted.get('targetObjectLoopNumber')
            targetObjectType2nd     = targetObjectExtracted.get('targetObjectType2nd')
            
            
            if targetObjectType and targetObjectLoopNumber:
                # get type of component tag and 1st block near to it.
                # typeTag = type of component like Ball Valve with Pneumatic actuator, pH metro, etc.
                typeTag,distance,near_block_found = findTypeBlockFromTag(component,components)
                # get description and connection type in the tag/sticker
                description = stripField(  attributes.get('DESCRIPTION') ) 
                connection_type = stripField( attributes.get('CONNECTIONTYPE') )
                
                # if description and type of componente not present then fetch description 
                # and type from the nearest block
                block_def_found = None
                entity_found = None
                if near_block_found:
                    entity_found = near_block_found.get("entity")
                    block_def  = near_block_found.get("block_def")
                    attributes_near_block_found = near_block_found.get('attributes')
                    if attributes_near_block_found:
                        if not(description):
                            description = attributes_near_block_found.get('DESCRIPTION')
                        if not(connection_type):
                            connection_type = attributes_near_block_found.get('CONNECTIONTYPE')
                            
                
                if description is None: description = ''
                if connection_type is None: connection_type= ''
                if typeTag is None: typeTag= ''
                # "tag_entity": component - the ezdxf.entities.insert.Insert obj of the tag/sticker
                # "target_entity": entity_found - the ezdxf.entities.insert.Insert obj which tag/sticker is pointing to,
                #                               if a sticker has 
                new_component_coded = {
                        "count":number,
                        'targetObjectType':targetObjectType,
                        'targetObjectLoopNumber':targetObjectLoopNumber,
                        'targetObjectType2nd':targetObjectType2nd,
                        'P&ID TAG': str(targetObjectType)+str(targetObjectLoopNumber)+str(targetObjectType2nd),
                        'TYPE':typeTag,
                        'DESCRIPTION':description,
                        'CONNECTIONTYPE':connection_type,
                        'insert_point': component.get('insert_point'),
                        'dimensions': component.get('dimensions'),
                        #"target_block_def" : block_def_found,
                        "target_entity": entity_found,
                        "tag_entity": component,
                        }
                if entity_found:
                    entity_to_be_extracted = entity_found
                else: entity_to_be_extracted = component['entity']
                
                for att in  entity_to_be_extracted.attribs:
                    if not(att.dxf.tag in new_component_coded.keys()):
                        new_component_coded.update({att.dxf.tag : att.dxf.text  })
                        
                
                bom.update( {number:new_component_coded})
            else:
                block_name = component['block_name']
                print(f"Warning: Block {block_name} not Valid")
                blocks_notValid.update({number:component})

                
    return bom, blocks_notValid

def print_bom(bom):
    print("\nGenerated BOM:\n")
    print(f"{'#':<3}|{'L':<4}|{'N':<5}|{'D':<4}|{'P&ID TAG':<10}|{'Type':<30}|{'Description':<30}")
    print("-" * 40)
    for it in bom.keys():
        component = bom[it]
        L = component['targetObjectType']
        N = component['targetObjectLoopNumber']
        D = component['targetObjectType2nd']
        pid_TAG = str(L)+str(N)+str(D)
        comp_type = component['TYPE']
        description=component['DESCRIPTION']
        if description == None: description=''
        if comp_type == None: comp_type=''
        print(f"{it:<3}|{L:<4}|{N:<5}|{D:<4}|{pid_TAG:<10}|{comp_type:<30}|{description:<30}")


# def sortingBOM_dict(bom_dxf,bytag = '' ):
#     bom_dxf 
    
#     # sort_key_for_pid_tag(tag)
#     bom_dxf_items = list(bom_dxf.items())
        
#     bom_dxf_items_sorted = sorted(bom_dxf_items, key=lambda x: x[1].get(bytag, ''))
    
#     sorted_bom = {}
#     for i_ in range(len(bom_dxf_items_sorted)):
#         i = i_+1
#         item = bom_dxf_items_sorted[i_][1]
#         item['count'] = i
#         sorted_bom[i] = item
#     return sorted_bom
    
def sort_bom_by_pid_tag(bom_dict):
    """
    Sort the BOM dictionary by the 'P&ID TAG' field using sort_key_for_pid_tag.
    
    Args:
        bom_dict (dict): A dictionary where values are row dictionaries containing
                         a "P&ID TAG" field.
                         
    Returns:
        dict: A new dictionary with entries sorted by the 'P&ID TAG'.
              (Python 3.7+ dictionaries preserve insertion order.)
    """
    sorted_items = sorted(
        bom_dict.items(),
        key=lambda item: sort_key_for_pid_tag(item[1].get("P&ID TAG", ""))
    )
    return {key: value for key, value in sorted_items}
    
def export_bom_to_excel(bom_data, template_path, output_path, highlight_duplicate=False):
    # Load the template file
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active  # Assuming the BOM sheet is the active one
    
    # # Mapping between BOM keys and Excel column headers
    print( header_mapping )

    
    # Find the corresponding column index for each header in the Excel template
    header_row = 1  # Assuming headers are in the first row
    column_indexes = {}
    
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=header_row, column=col).value
        for bom_key, header_name in header_mapping.items():
            if header_value == header_name:
                column_indexes[bom_key] = col
        if header_value == 'P&ID TAG':
            pid_tag_col = col  # Find the column for "P&ID TAG"
    
    # Start writing BOM data after the header row (e.g., starting at row 2)
    start_row = 2
    for i, bom_item in enumerate(bom_data.values(), start=start_row):
        # Write the regular BOM fields
        for bom_key, value in bom_item.items():
            if bom_key in column_indexes:
                col = column_indexes[bom_key]
                sheet.cell(row=i, column=col).value = value
        
        # Generate and write the "P&ID TAG" (L + N + D)
        l_value = bom_item.get('targetObjectType', '')
        n_value = str(bom_item.get('targetObjectLoopNumber', ''))
        d_value = bom_item.get('targetObjectType2nd', '')
        
        pid_tag = f"{l_value}{n_value}{d_value}"
        sheet.cell(row=i, column=pid_tag_col).value = pid_tag
    
    sort_rows_by_pid_tag(sheet)
    # Trigger duplicate highlighting if the flag is true
    if highlight_duplicate:
         print("\nHighlighting Duplicate 'P&ID TAG' Values in Purple:")
         highlight_duplicate_tags_in_excel(sheet, 'P&ID TAG', color="800080")  # Purple
    # Save the workbook with a new name
    workbook.save(output_path)
    print(f"BOM successfully exported to {output_path}")

def extract_bom_from_dxf(dwg_file):
    components, docDxf = extract_blocks_with_attributes_and_dimensions(dwg_file)
    
    # filter only the blocks stickers TAGs to determine the components in BOM
    bom, blocks_notValid = generate_bom(components)
    bom = sort_bom_by_pid_tag(bom)
    #print(bom)
    # Print the formatted BOM with dimensions
    print_bom(bom)
    return bom,docDxf
    

# def load_bom_from_excel(file_path):
#     """ Load BOM data from Excel file. """
#     df = pd.read_excel(file_path)
    
#     # Assuming columns like 'L', 'N', 'D', 'Type', 'Description'
#     df['N'] = df['N'].fillna(0).astype(int).astype(str)  # Convert N to int, then to string
#     df['D'] = df['D'].fillna('').astype(str)  # Convert D to empty string if NaN
    
#     # Construct P&ID TAG by concatenating L, N, and D
#     df['P&ID TAG'] = df['L'].astype(str) + df['N'] + df['D']
    
#     return df

def load_bom_from_excel_to_JSON(file_path):
    """ Load BOM data from Excel file and return it as a JSON-like dictionary with line numbers. """
    
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Assumes data is in the active sheet
    
    # Initialize a dictionary to hold the BOM data with line numbers
    bom_data = {}
    
    # Get headers (first row)
    headers = [cell.value for cell in sheet[1]]
    
    # Iterate over rows, starting from the second row (assuming the first row is the header)
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        
        if all(cell is None or cell == '' or cell == 0 for cell in row):
            continue  # Skip this row if it is empty
        row_data = {headers[j]: (row[j] if row[j] is not None else '') for j in range(len(headers))}
        
        # Convert 'N' to string after replacing NaN (or empty) values with 0
        row_data['N'] = str(row_data['N']) if row_data['N'] != '' else ''
        
        # Convert 'D' to an empty string if NaN or None
        row_data['D'] = str(row_data['D']) if row_data['D'] != '' else ''
        
        # Construct P&ID TAG by concatenating L, N, and D
        row_data['P&ID TAG'] = str(row_data['L']) + row_data['N'] + row_data['D']
        
        # Add the row's data to the bom_data dictionary with the row number as the key
        bom_data[i] = row_data
    
    
    #bom_data = sortingBOM_dict(bom_data,bytag = 'P&ID TAG' )
    bom_data = sort_bom_by_pid_tag(bom_data)
    workbook.close()
    # Return data in JSON-like dictionary format with line numbers as keys
    return bom_data

# def convert_bom_dxf_to_dataframe(bom_dxf):
#     """ Convert BOM dictionary from DXF to a pandas DataFrame. """
#     if not isinstance(bom_dxf, dict):
#         raise ValueError("Input must be a dictionary.")
        
#     # Create a list of dictionaries for each item in the BOM
#     bom_list = []
#     for item in bom_dxf.values():
#         bom_list.append({
#             '#': item['count'],
#             'L': item['targetObjectType'],
#             'N': item['targetObjectLoopNumber'],
#             'D': item['targetObjectType2nd'],
#             'Type': item['TYPE'],
#             'Description': item['DESCRIPTION']
#         })
    
#     # Convert the list to a DataFrame
#     bom_df = pd.DataFrame(bom_list)

#     # Create the P&ID TAG column
#     bom_df['P&ID TAG'] = (
#         bom_df['L'].astype(str) +
#         bom_df['N'].astype(str) +
#         bom_df['D'].astype(str)
#     )

#     return bom_df


def convert_bom_dxf_to_JSON(bom_dxf):
    """ Convert BOM dictionary from DXF changing keys names as the headers in excell. """
    if not isinstance(bom_dxf, dict):
        raise ValueError("Input must be a dictionary.")
    
    bom_dxf2 = {}
    for i in bom_dxf.keys():
        comp = bom_dxf[i].copy()
        for key in header_mapping.keys():
            if key in comp.keys():
                key_new = header_mapping[key]
                value = comp[key]
                comp.update({key_new:value})
                del comp[key]
        L = comp['L']
        N = comp['N']
        D = comp['D']
        PID_TAG = str(L)+str(N)+str(D)
        comp.update({'P&ID TAG':PID_TAG,  })
        bom_dxf2[i] = comp

    return bom_dxf2

# def highlight_missing_item_in_excel_old(item, sheet,bom_revised):
#     # Find the row in the revised BOM DataFrame
#     row = bom_revised[bom_revised['P&ID TAG'] == item]
#     if not row.empty:
#         row_index = row.index[0] + 2  # Adjust for header row
#         # Highlight the cells in red
#         for col in range(1, sheet.max_column + 1):  # Adjust to the range of your actual columns
#             sheet.cell(row=row_index, column=col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        
# def add_missing_items_to_excel_old(missing_items, sheet, bom_dxf_df):
#     """ Add missing items from the DXF BOM to the Excel sheet and highlight them in grey. """
    
#     # Create a mapping of column headers to their indices
#     header_mapping = {sheet.cell(row=1, column=col).value: col for col in range(1, sheet.max_column + 1)}
#     header_length = len(header_mapping.keys()) 
#     value_to_write = ['#','L','N','D','P&ID TAG','Type','Description']
#     # Iterate through the missing items
#     for item in missing_items:
#         dxf_row = bom_dxf_df[bom_dxf_df['P&ID TAG'] == item].iloc[0]
        
#         # # Create a new row based on the header mapping
#         new_row = ['' for k_ in range(header_length)]
#         for col_val in value_to_write:
#             value_row = dxf_row[col_val] 
#             pos_col = header_mapping[col_val]-1
#             new_row[pos_col]=value_row
        
        
        
#         # Find the next empty row in the sheet
#         for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
#             if all(cell.value is None for cell in sheet[row]):
#                 # Replace the empty row with the new row
#                 for col_header, value in zip(header_mapping.keys(), new_row):
#                     col_index = header_mapping[col_header]
#                     sheet.cell(row=row, column=col_index).value = value
                
#                 # Highlight the new row in grey
#                 for col in range(1, len(new_row) + 1):
#                     sheet.cell(row=row, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
#                 break
#         else:
#             # If no empty row found, append to the end of the sheet
#             start_row = sheet.max_row + 1
#             for col_header, value in zip(header_mapping.keys(), new_row):
#                 col_index = header_mapping[col_header]
#                 sheet.cell(row=start_row, column=col_index).value = value
            
#             # Highlight the new row in grey
#             for col in range(1, len(new_row) + 1):
#                 sheet.cell(row=start_row, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")


# def compare_boms(bom_dxf_df, bom_revised, revised_excel_file=None, highlight_missing=False, import_missingDXF2BOM = False):
#     """ Compare two BOMs and print differences. Optionally highlight missing components in red. """
    
#     sheet = None
#     if highlight_missing and revised_excel_file:
#         workbook = openpyxl.load_workbook(revised_excel_file)
#         sheet = workbook.active
    
#     # Filter the revised BOM similarly
#     bom_revised = bom_revised.dropna(subset=['L', 'N', 'D', 'P&ID TAG'])
    
#     # Create sets of P&ID TAGs for comparison
#     bom_dxf_set = set(bom_dxf_df['P&ID TAG'])
#     bom_revised_set = set(bom_revised['P&ID TAG'])

#     # Components in BOM_dxf but not in BOM_revised
#     missing_in_revised = bom_dxf_set - bom_revised_set
#     if missing_in_revised:
#         print("\nComponents in DXF but not in the revised BOM:")
#         for item in missing_in_revised:
#             print(item)

#     # Components in BOM_revised but not in BOM_dxf
#     missing_in_dxf = bom_revised_set - bom_dxf_set
#     if missing_in_dxf:
#         print("\nComponents in revised BOM but not in DXF BOM:")
#         for item in missing_in_dxf:
#             print(item)

#             # Highlight in the revised Excel file if specified
#             if sheet:
#                 #print('red')
#                 highlight_missing_item_in_excel(item, sheet, bom_revised)
#                 #workbook.save(revised_excel_file)



#     # New feature: Ask if the user wants to import missing items from the DXF
#     if missing_in_revised:
        
#         #if import_missing == 'yes' and sheet:
#         if import_missingDXF2BOM and sheet: 
#             add_missing_items_to_excel(missing_in_revised, sheet, bom_dxf_df)
#             #workbook.save(revised_excel_file)  # Save the updated Excel file
            
#     if sheet:
#         workbook.save(revised_excel_file)  # Save changes to the Excel file
#     # Compare attributes
#     print("\nComparing attributes of matching components...")
#     for tag in bom_dxf_set.intersection(bom_revised_set):
#         dxf_row = bom_dxf_df[bom_dxf_df['P&ID TAG'] == tag].iloc[0]
#         revised_row = bom_revised[bom_revised['P&ID TAG'] == tag].iloc[0]
        
#         # Check for differences in attributes (Type, Description, etc.)
#         for column in ['Type', 'Description']:
#             dxf_value = dxf_row[column] if pd.notna(dxf_row[column]) else None
#             if dxf_value == '': dxf_value = None
#             revised_value = revised_row[column] if pd.notna(revised_row[column]) else None
            
#             if dxf_value != revised_value:
#                 print(f"Component {tag}: {column} differs. DXF: {dxf_value}, Revised: {revised_value}")
    
#     return missing_in_revised, missing_in_dxf

def find_duplicates(bom_dict, field):
    """
    Find duplicate rows in bom_dict based on a given field.

    Args:
        bom_dict (dict): A dictionary where keys are row identifiers and values are dictionaries.
        field (str): The field name to check duplicates on (e.g., "P&ID TAG").

    Returns:
        dict: A dictionary mapping each duplicate field value to a list of row IDs (keys) that share that value.
              Only field values that occur more than once are included.
    """
    seen = {}
    duplicates = {}
    
    for row_id, row in bom_dict.items():
        value = row.get(field)
        if value in seen:
            # First duplicate occurrence: add the already-seen row.
            if value not in duplicates:
                duplicates[value] = [seen[value]]
            duplicates[value].append(row_id)
        else:
            seen[value] = row_id

    return duplicates

def make_color_mapping(colour_mapping ,list_tag,color):
    if not(isinstance(colour_mapping, dict)):
        colour_mapping={}
    if list_tag:
        list_tag_list = list(list_tag)
        for key_tag in list_tag_list:
            colour_mapping.update({key_tag:color })
        return colour_mapping
    else: 
        return {}
    
    
# Filter bom_dxf and bom_revised to include only entries with 'L', 'N', 'D', 'P&ID TAG' present and not empty
def filter_bom(bom):
    filtered_bom = {}
    for key, comp in bom.items():
        if all(comp.get(field) not in [None, ''] for field in ['L', 'N', 'P&ID TAG']):
            filtered_bom[key] = comp
    return filtered_bom

def compare_bomsJSON(
    bom_dxf, 
    bom_revisedJSON, 
    # highlight_duplicate=False, 
    # highlight_missing=False, 
    # import_missingDXF2BOM=False
    ):
    """
    Compares two BOMs and optionally highlights duplicates and missing components.

    Args:
        bom_dxf: BOM extracted from the DXF file.
        bom_revisedJSON: Revised BOM in JSON format.
        revised_excel_file: Path to the revised Excel file (optional).
        highlight_duplicate: Flag to trigger highlighting of duplicate P&ID tags.
        highlight_missing: Flag to trigger highlighting of missing components.
        import_missingDXF2BOM: Flag to trigger import of missing components.
    """
    # sheet = None
    # if (highlight_missing or import_missingDXF2BOM) and revised_excel_file:
    #     workbook = openpyxl.load_workbook(revised_excel_file)
    #     sheet = workbook.active
        
    bom_dxf_filtered = filter_bom(bom_dxf)
    bom_revised_filtered = filter_bom(bom_revisedJSON)
    

    # Create sets of 'P&ID TAG's
    bom_dxf_set = set(comp['P&ID TAG'] for comp in bom_dxf_filtered.values())
    bom_revised_set = set(comp['P&ID TAG'] for comp in bom_revised_filtered.values())

    # Components in BOM_dxf but not in BOM_revised
    missing_in_revised = bom_dxf_set - bom_revised_set
    if missing_in_revised:
        print("\nComponents in DXF but not in the revised BOM:")
        for item in missing_in_revised:
            print(item)

    # Components in BOM_revised but not in BOM_dxf
    missing_in_dxf = bom_revised_set - bom_dxf_set
    if missing_in_dxf:
        print("\nComponents in revised BOM but not in DXF BOM:")
        for item in missing_in_dxf:
            print(item)
            # Highlight in the revised Excel file if specified
            # if sheet and highlight_missing:
            #     highlight_missing_item_in_excel(item, sheet)
    

    # Compare attributes of matching components
    print("\nComparing attributes of matching components...")
    for tag in bom_dxf_set.intersection(bom_revised_set):
        # Find the component in each BOM
        dxf_comp = next(comp for comp in bom_dxf_filtered.values() if comp['P&ID TAG'] == tag)
        revised_comp = next(comp for comp in bom_revised_filtered.values() if comp['P&ID TAG'] == tag)
        # Compare 'Type' and 'Description'
        for column in ['Type', 'Description']:
            dxf_value = dxf_comp.get(column, None) or None
            revised_value = revised_comp.get(column, None) or None
            if dxf_value != revised_value:
                print(f"Component {tag}: {column} differs. DXF: {dxf_value}, Revised: {revised_value}")
                
    return missing_in_revised, missing_in_dxf

def update_XLS_add_missing_items_highlight(
        workbook_xls,
        bom_dxf,
        bom_revisedJSON,
        missing_in_revised, 
        missing_in_dxf,
        highlight_duplicate=False, 
        highlight_missing=False,  
        import_missingDXF2BOM=False,
        missing_in_dxf_color="FF0000",         # Default red
        missing_in_revised_color="CCCCCC",      # Default gray
        duplicate_color="800080"                # Default purple
    ):
    """
    Updates the Excel workbook by adding missing items and highlighting components based on various criteria.

    Args:
        workbook_xls: An openpyxl workbook object.
        bom_dxf: Dictionary of DXF BOM data.
        bom_revisedJSON: Dictionary of revised BOM data (JSON structure).
        missing_in_revised: A list of items missing in revised BOM.
        missing_in_dxf: A list of items missing in DXF BOM.
        highlight_duplicate (bool): Flag to highlight duplicate P&ID TAG values.
        highlight_missing (bool): Flag to highlight missing components.
        import_missingDXF2BOM (bool): Flag that indicates whether to import missing DXF items.
        missing_in_dxf_color (str): Hex color value for highlighting missing items in DXF.
        missing_in_revised_color (str): Hex color value for highlighting missing items in revised BOM.
        duplicate_color (str): Hex color value for highlighting duplicate P&ID TAG values.

    Returns:
        The updated workbook_xls.
    """
                                                
    #workbook = openpyxl.load_workbook(revised_excel_file)
    sheet = workbook_xls.active
    
    bom_dxf_filtered = filter_bom(bom_dxf)
    
    # Add missing items from DXF to Excel if requested
    if missing_in_revised and import_missingDXF2BOM and sheet:
        add_missing_items_to_excel(missing_in_revised, sheet, bom_dxf_filtered)
    
    
    # sorting
    sort_rows_by_pid_tag(sheet)
    
    
    # highlight
    if missing_in_dxf and highlight_missing and  sheet:
        print("\nHighlighting Components in revised BOM but not in DXF BOM:")
        for item in missing_in_dxf:
            highlight_missing_item_in_excel(item, sheet,color = missing_in_dxf_color) # Red color RGB
    if missing_in_revised and import_missingDXF2BOM and sheet and highlight_missing:
        print("\nHighlighting Components in DXF but not in the revised BOM:")
        for item in missing_in_revised:
            highlight_missing_item_in_excel(item, sheet,color = missing_in_revised_color) # Gray color RGB

   # Trigger duplicate highlighting if the flag is true
    if highlight_duplicate and sheet:
        print("\nHighlighting Duplicate 'P&ID TAG' Values in Purple:")
        highlight_duplicate_tags_in_excel(sheet, 'P&ID TAG', color=duplicate_color)  # Purple
        
    return workbook_xls



def highlight_missing_item_in_excel(item, sheet,color = "FF0000"):
    """Highlight the row corresponding to 'item' in the Excel sheet."""
    # Assuming the first row is the header
    pid_tag_col = None
    # Find the column index for 'P&ID TAG'
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        if header_value == 'P&ID TAG':
            pid_tag_col = col
            break
    if pid_tag_col is None:
        print("P&ID TAG column not found in Excel sheet.")
        return
    # Iterate over the rows to find the item
    for row in range(2, sheet.max_row + 1):  # Start from 2 to skip header
        cell_value = sheet.cell(row=row, column=pid_tag_col).value
        if cell_value == item:
            # Highlight the entire row in red
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            break  # Exit after highlighting the row


def highlight_duplicate_tags_in_excel(sheet, column_name, color="800080"):
    """
    Highlights cells in the given column with duplicate values.

    Args:
        sheet: The Excel sheet object.
        column_name: The name of the column to check for duplicates (e.g., 'P&ID TAG').
        color: The color code for highlighting duplicates (default: Purple '800080').
    """
    column_index = None
    header_row = list(sheet[1])  # Assuming the first row contains headers
    
    # Find the column index for the specified column name
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == column_name:
            column_index = idx
            break

    if column_index is None:
        print(f"Column '{column_name}' not found in the sheet.")
        return

    # Track duplicates
    tag_counts = {}
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
        tag_value = row[0].value
        if tag_value:
            if tag_value in tag_counts:
                tag_counts[tag_value].append(row[0])
            else:
                tag_counts[tag_value] = [row[0]]

    # Highlight duplicates
    for tag, cells in tag_counts.items():
        if len(cells) > 1:  # Only highlight if more than one occurrence
            for cell in cells:
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")
                print(f"Highlighted duplicate tag '{tag}' in cell {cell.coordinate}.")
     
def add_missing_items_to_excel(missing_items, sheet, bom_dxf):
    """
    Add missing items from the DXF BOM to the Excel sheet and highlight them (e.g. in grey).
    
    Optimizations:
      - Pre-build a lookup for bom_dxf by 'P&ID TAG' for O(1) access.
      - Scan the sheet once to determine which rows are empty.
    """
    # Pre-build a mapping from P&ID TAG to its component dictionary.
    dxf_lookup = {comp['P&ID TAG']: comp for comp in bom_dxf.values() if comp.get('P&ID TAG')}
    
    # Create a mapping of column headers to their indices.
    header_mapping = {
        sheet.cell(row=1, column=col).value: col 
        for col in range(1, sheet.max_column + 1)
    }
    
    # Helper function to check if a row is empty.
    # (We assume that a row is empty if all its cells are None or the empty string.)
    def is_row_empty(row):
        return all(cell.value in (None, '') for cell in row)
    
    # Pre-calculate a list of empty row numbers starting from row 2.
    empty_rows = [
        row_num for row_num in range(2, sheet.max_row + 1)
        if is_row_empty(sheet[row_num])
    ]
    
    # Process each missing item.
    for item in missing_items:
        # Retrieve the corresponding component from bom_dxf using the lookup.
        dxf_comp = dxf_lookup.get(item)
        if not dxf_comp:
            # If not found, you might choose to log or continue.
            continue
        
        # Determine which row to use: if an empty row exists, pop it; otherwise, append after the last row.
        if empty_rows:
            row_num = empty_rows.pop(0)
        else:
            row_num = sheet.max_row + 1
        
        # Write each header's value from the DXF component into the sheet.
        for header, col_index in header_mapping.items():
            value = dxf_comp.get(header, '')
            sheet.cell(row=row_num, column=col_index).value = value

# def add_missing_items_to_excel_old(missing_items, sheet, bom_dxf):
#     """Add missing items from the DXF BOM to the Excel sheet and highlight them in grey."""
#     # Create a mapping of column headers to their indices
#     header_mapping_xls = {sheet.cell(row=1, column=col).value: col for col in range(1, sheet.max_column + 1)}

#     # Helper function to check if a row is empty
#     def is_row_empty(row):
#         return all(cell.value in (None, '') for cell in row)

#     # Iterate through the missing items
#     for item in missing_items:
#         # Find the component in bom_dxf
#         dxf_comp = next(comp for comp in bom_dxf.values() if comp['P&ID TAG'] == item)

#         # Search for the first available empty row starting from row 2 (assuming row 1 has headers)
#         row_num = None
#         for row in range(2, sheet.max_row + 1):
#             if is_row_empty(sheet[row]):
#                 row_num = row
#                 break

#         # If no empty row was found, append to the end of the sheet
#         if row_num is None:
#             row_num = sheet.max_row + 1

#         # Write data to the identified row (either an existing empty row or a new row)
#         for header, col_index in header_mapping_xls.items():
#             value = dxf_comp.get(header, '')
#             sheet.cell(row=row_num, column=col_index).value = value

#             # Highlight the new row in grey
#             #sheet.cell(row=row_num, column=col_index).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# def sort_rows_by_pid_tag(sheet):
#     """Sort rows in the Excel sheet by 'P&ID TAG' column, ignoring empty rows and renumbering the '#' column."""
    
#     # Find the column index for 'P&ID TAG' and '#' columns
#     pid_tag_col = None
#     num_col = None  # Column for '#'
    
#     for col in range(1, sheet.max_column + 1):
#         header_value = sheet.cell(row=1, column=col).value
#         if header_value == 'P&ID TAG':
#             pid_tag_col = col
#         elif header_value == '#':
#             num_col = col
#         if pid_tag_col is not None and num_col is not None:
#             break
    
#     if pid_tag_col is None:
#         print("P&ID TAG column not found in Excel sheet.")
#         return

#     if num_col is None:
#         print("'#' column not found in Excel sheet.")
#         return

#     # Extract all non-empty rows (excluding the header row)
#     rows = []
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         # Check if the row is fully empty
#         if any(cell is not None and cell != '' for cell in row):
#             rows.append(row)

#     # Sort rows based on the 'P&ID TAG' value (using the pid_tag_col - 1 to match zero-based index)
#     sorted_rows = sorted(rows, key=lambda x: (x[pid_tag_col - 1] or '').lower())

#     # Clear existing rows (excluding the header)
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#         for cell in row:
#             cell.value = None

#     # Write the sorted rows back to the sheet and renumber the '#' column
#     for i, row_data in enumerate(sorted_rows, start=2):
#         for j, value in enumerate(row_data, start=1):
#             # Renumber the '#' column
#             if j == num_col:
#                 sheet.cell(row=i, column=j).value = i - 1  # Renumber from 1
#             else:
#                 sheet.cell(row=i, column=j).value = value


    
def sort_key_for_pid_tag(tag):
    """
    Generate a sorting key from a P&ID TAG string.

    The key is a tuple: (prefix, number, suffix) where:
      - prefix is the initial alphabetic part in lowercase,
      - number is the numeric part parsed as an integer,
      - suffix is the remaining alphabetic part in lowercase.
      
    If the tag doesn't match the expected pattern, it returns a key
    that sorts the tag at the end.
    """
    tag=str(tag)
    pattern = re.compile(r"^([A-Za-z]+)(\d+)([A-Za-z]*)$")
    match = pattern.match(tag or "")
    if match:
        prefix, num_str, suffix = match.groups()
        return (prefix.lower(), int(num_str), suffix.lower())
    else:
        # Return a triple that always sorts after valid tags.
        # "{" comes after "z" in ASCII, and float('inf') ensures numeric value is highest.
        #return (float('inf'), tag.lower())
        #return (0, tag.lower(), 0)
        return ('@',-float('inf'), tag.lower())

def sort_rows_by_pid_tag(sheet):
    """Sort rows in the Excel sheet by 'P&ID TAG' using a custom sort key, ignoring empty rows, and renumber the '#' column."""
    
    pid_tag_col = None
    num_col = None  # Column for '#'
    
    # Find the column indices for 'P&ID TAG' and '#' (numbering) in header row (assumed to be row 1)
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        if header_value == 'P&ID TAG':
            pid_tag_col = col
        elif header_value == '#':
            num_col = col
        if pid_tag_col is not None and num_col is not None:
            break

    if pid_tag_col is None:
        print("P&ID TAG column not found in Excel sheet.")
        return

    if num_col is None:
        print("'#' column not found in Excel sheet.")
        return

    # Gather all non-empty rows (excluding the header) as tuples of cell values.
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Consider row non-empty if any cell is not None and not empty
        if any(cell is not None and cell != '' for cell in row):
            rows.append(row)

    # Sort rows using our custom sort key for the 'P&ID TAG' column.
    # Adjust column index (Python uses zero-based indexing)
    sorted_rows = sorted(rows, key=lambda r: sort_key_for_pid_tag(r[pid_tag_col - 1]))

    # Clear existing rows below header.
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Write the sorted rows back into the sheet and renumber the '#' column.
    for i, row_data in enumerate(sorted_rows, start=2):
        for j, value in enumerate(row_data, start=1):
            if j == num_col:
                sheet.cell(row=i, column=j).value = i - 1  # Renumber '#' column starting from 1.
            else:
                sheet.cell(row=i, column=j).value = value